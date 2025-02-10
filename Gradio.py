import gradio as gr
import requests
import json
import csv
import pandas as pd
import spacy
import openpyxl #for Excel file
from openpyxl.styles import Font, Alignment
from bs4 import BeautifulSoup
import re

nlp = spacy.load("en_core_web_sm")

# Ollama Docker server details
OLLAMA_SERVER_URL = "http://ollama:11434"  # It should work as it's on the same Docker network.

import json
import requests

def query_ollama(user_input):
    """
    Sends structured data to Ollama and returns the response.
    """
    url = f"{OLLAMA_SERVER_URL}/api/generate"

    #parse user input to prompt
    print(f"Received input: {user_input}")  # Debugging log
    match = re.search(r'The threat is:(.*?)\s*This threat is caused by:(.*?)\s*This puts the(.*?)at risk\s*Leading to consequences like:(.*)',user_input['text'],re.DOTALL)
    if match == None:
        return "wrong format, please refer to format provided"+user_input['text']
    Threat_Event=match.group(1)
    Vulnerability=match.group(2)
    Asset=match.group(3)
    Consequence=match.group(4)

    prompt = f"""
    The threat is: {Threat_Event}
    This threat is caused by: {Vulnerability}
    This puts the {Asset} at risk
    Leading to consequences like: {Consequence}
    """
    print(f"prompt recieved:{prompt}")

    payload = {
        "model": "unsloth_model",
        "prompt": prompt.strip()
    }

    try:
        print(f"Sending request to Ollama: {json.dumps(payload, indent=2)}")  # Debugging log
        response = requests.post(url, json=payload, timeout=60, stream=True)  # Set stream to True

        print(f"Response Status Code: {response.status_code}")  # Debugging log
        print(f"Raw Response: {response.text}")  # Print raw response content for debugging

        # Initialize an empty string to store the full response
        full_response = ""

        # Check if response content is empty before processing it
        if not response.text.strip():
            return "Error: Empty response from Ollama."

        # Read the streaming response content
        for chunk in response.iter_lines():
            if chunk:  # Skip empty chunks
                try:
                    chunk_data = json.loads(chunk.decode("utf-8"))
                    print(f"Chunk Response Data: {json.dumps(chunk_data, indent=2)}")  # Debugging log
                    full_response += chunk_data.get("response", "")
                except json.JSONDecodeError:
                    print(f"Error: Failed to decode chunk. Raw chunk: {chunk.decode('utf-8')}")
                    continue

        # Final check to make sure there's some content
        if full_response.strip():
            return full_response.strip()
        else:
            return "Error: No meaningful response from Ollama."

    except requests.exceptions.RequestException as e:
        return f"Error communicating with Ollama: {e}"

#----- Read uploaded CSV file for instructions ----------------------------------------
def parse_instruction_csv(f):
    query=[]
    result=''
    with open(f,'r') as f1:
        reader=csv.reader(f1)
        for box in reader:
            query.append(box[0].replace('\n','').split('>')) # Cleaning unwanted string and separate by stages 
    print(query)  # Debug: Log parsed instructions
    for line in query:
        for i in line:
            result += "i: "+query_ollama(i) + "\n"   # Send query to Ollama
    return result

# Receive uploaded file and send to llama
def upload_file(f):
    f_parsed = parse_instruction_csv(f)
    return f_parsed

# Export editable output to Excel
def editable_export_excel(editable_output):
    try:
        lines = editable_output.split('\n')
        wb = openpyxl.Workbook()
        ws = wb.active

        for i, line in enumerate(lines):
            doc = nlp(line)
            ws.cell(row=i+1, column=1).value = line
            for token in doc:
                if token.pos_ == "VERB":
                    ws.cell(row=i+1, column=1).font = Font(color='FFFF0000')  
                elif token.pos_ == "NOUN":
                    ws.cell(row=i+1, column=1).font = Font(color='FF0000FF')  
                elif token.text.lower() in ["achieve", "gain", "obtain"]:
                    ws.cell(row=i+1, column=1).font = Font(color='FF00FF00')  

        file_name = "output.xlsx"
        wb.save(file_name)
        return gr.File(file_name)
    except Exception as e:
        return str(e)

with gr.Blocks(theme=gr.themes.Citrus()) as demo:
    gr.Markdown("<center><h1>R.A.P - Risk Assessment Prompt</h1></center>")
    gr.Markdown("<center>This is a LLM semi-automation used for Risk Assessments. Outputs are color-coded for easy reference.<center>")

    with gr.Row():
        with gr.Column():
            input = gr.MultimodalTextbox(interactive=True, file_count="multiple", placeholder="Enter text or upload file")
            with gr.Row():
                clear_button = gr.Button("Clear")
                submit_button = gr.Button("Submit")
            gr.Markdown("<b>How to chat with R.A.P:<b>")
            gr.Markdown("The threat is: {Threat_Event}<br>This threat is caused by: {Vulnerability}<br>This puts the {Asset} at risk<br>Leading to consequences like: {Consequence}")
        with gr.Column():
            output = gr.Textbox(label="Editable Output", interactive=True)
            export_button = gr.Button("Export Output - Excel")
            export_file = gr.File(label="Exported Edited Excel File")
            export_button.click(editable_export_excel, inputs=output, outputs=export_file)

        def prompt(question):
            print(f"Received input question: {question}")  # Debug: Log the user input
            ollama_response = query_ollama(question)
            return ollama_response

        def clear_input():
            return ""

        clear_button.click(clear_input, inputs=None, outputs=input)
        submit_button.click(prompt, inputs=input, outputs=output)

# Launch the app
demo.launch(server_name="0.0.0.0", server_port=7860)

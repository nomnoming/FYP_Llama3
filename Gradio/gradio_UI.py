import gradio as gr
import requests
import json
import csv
import pandas as pd
import spacy
import openpyxl
from openpyxl.styles import Font, Alignment
from bs4 import BeautifulSoup
from input_parser import parse_input

nlp = spacy.load("en_core_web_sm")

# Ollama Docker server details
OLLAMA_SERVER_URL = "http://ollama:11434"  # It should work as it's on the same Docker network.

def query_ollama(parsed_input):
    """
    Sends structured data to Ollama and returns the full, assembled response.
    Handles fragmented responses from Ollama with enhanced error handling.
    """
    url = f"{OLLAMA_SERVER_URL}/api/generate"

    prompt = f"""
    The threat is: {parsed_input.get('threat', 'Unknown')}
    This threat is caused by: {parsed_input.get('cause', 'Unknown')}
    This puts the {parsed_input.get('risk', 'Unknown')} at risk
    Leading to consequences like: {parsed_input.get('consequences', 'Unknown')}
    """

    payload = {
        "model": "unsloth_model",
        "prompt": prompt.strip()
    }

    try:
        print(f"Sending request to Ollama: {json.dumps(payload, indent=2)}")  # Debugging log
        response = requests.post(url, json=payload, timeout=120, stream=True)  # Increased timeout to 120s

        print(f"Response Status Code: {response.status_code}")  # Debugging log
        if response.status_code != 200:
            print(f"Error: Ollama returned status code {response.status_code}")  # Log response code if not OK
            return f"Error: Ollama returned status code {response.status_code}"

        print(f"Raw Response: {response.text}")  # Debugging log

        # Initialize an empty string to store the full response
        full_response = ""

        # Read the streaming response content
        for chunk in response.iter_lines():
            if chunk:  # Skip empty chunks
                try:
                    chunk_data = json.loads(chunk.decode("utf-8"))
                    print(f"Chunk Response Data: {json.dumps(chunk_data, indent=2)}")  # Debugging log
                    if 'response' in chunk_data:
                        full_response += chunk_data['response']
                    else:
                        print(f"Warning: Missing 'response' field in chunk: {chunk_data}")
                except json.JSONDecodeError as e:
                    print(f"Error: Failed to decode chunk. Raw chunk: {chunk.decode('utf-8')}, Error: {str(e)}")
                    continue

        if full_response.strip():
            print(f"Final response: {full_response}")  # Debugging log
            return full_response.strip()
        else:
            print("Error: No meaningful response from Ollama after assembling chunks.")  # Debugging log
            return "Error: No meaningful response from Ollama."

    except requests.exceptions.RequestException as e:
        print(f"Request Exception: {e}")  # Log the error message for better debugging
        return f"Error communicating with Ollama: {e}"


#------ handling the input and passing it through an input parser -----------------------
def handle_gradio_input(user_input):
    """
    Parses the Gradio input, extracts structured information, and sends it to Ollama.
    """
    print(f"Received input: {user_input}")  # Debugging log
    parsed_data = parse_input(user_input)

    if "error" in parsed_data:
        return parsed_data["error"]  # Return error message if parsing fails

    print(f"Parsed data: {json.dumps(parsed_data, indent=2)}")  # Debugging log
    return query_ollama(parsed_data) # sends it to the ollama query on top ^


#----- Read uploaded CSV file for instructions ----------------------------------------
def parse_instruction_csv(f):
    query=[]
    result=''
    with open(f,'r') as f1:
        reader=csv.reader(f1)
        for box in reader:
            query.append(box[0].replace('\n','').split('>')) # Cleaning unwanted string and separate by stages 
    print(query)  # Debug: Log parsed instructions
    for line in query:
        for i in line:
            result += "i: "+query_ollama(i) + "\n"   # Send query to Ollama
    return result

# Receive uploaded file and send to llama
def upload_file(f):
    f_parsed = parse_instruction_csv(f)
    return f_parsed

# Export editable output to Excel
def editable_export_excel(editable_output):
    try:
        lines = editable_output.split('\n')
        wb = openpyxl.Workbook()
        ws = wb.active

        for i, line in enumerate(lines):
            doc = nlp(line)
            ws.cell(row=i+1, column=1).value = line
            for token in doc:
                if token.pos_ == "VERB":
                    ws.cell(row=i+1, column=1).font = Font(color='FFFF0000')  
                elif token.pos_ == "NOUN":
                    ws.cell(row=i+1, column=1).font = Font(color='FF0000FF')  
                elif token.text.lower() in ["achieve", "gain", "obtain"]:
                    ws.cell(row=i+1, column=1).font = Font(color='FF00FF00')  

        file_name = "output.xlsx"
        wb.save(file_name)
        return gr.File(file_name)
    except Exception as e:
        return str(e)

with gr.Blocks(theme=gr.themes.Citrus()) as demo:
    gr.Markdown("<center><h1>R.A.P - Risk Assessment Prompt</h1></center>")
    gr.Markdown("<center>This is a LLM semi-automation used for Risk Assessments. Outputs are color-coded for easy reference.<center>")

    with gr.Row():
        with gr.Column():
            input = gr.Textbox(label="Enter the Risk Assessment Prompt", placeholder="Enter threat, cause, risk, and consequences in the required format.")
            with gr.Row():
                clear_button = gr.Button("Clear")
                submit_button = gr.Button("Submit")
            gr.Markdown("<b>How to chat with R.A.P:</b>")
            gr.Markdown("The threat is: {Threat_Event}<br>This threat is caused by: {Vulnerability}<br>This puts the {Asset} at risk<br>Leading to consequences like: {Consequence}")
        with gr.Column():
            output = gr.Textbox(label="Ollama Response", interactive=True)
            export_button = gr.Button("Export Output - Excel")
            export_file = gr.File(label="Exported Edited Excel File")
            export_button.click(editable_export_excel, inputs=output, outputs=export_file)

        def prompt(question):
            print(f"Received input question: {question}")  # Debug: Log the user input
            ollama_response = handle_gradio_input(question)  # Send input through the parser and Ollama query
            return ollama_response

        def clear_input():
            return ""

        clear_button.click(clear_input, inputs=None, outputs=input)
        submit_button.click(prompt, inputs=input, outputs=output)

# Launch the app
demo.launch(server_name="0.0.0.0", server_port=7860)
